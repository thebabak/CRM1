# core/views.py
from __future__ import annotations

from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST, require_http_methods
from django.db import transaction
from django.db.models import Q
from django.db.models.functions import TruncDate
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .forms import RequestForm, LeaveRequestForm
from .models import (
    Request,
    RequestStatus,
    Decision,
    Role,
    apply_decision,
    can_user_approve,
    leave_year_stats,
    Profile,
    LeaveRequest,
    Approval,
)

# Optional registry for the "New Request" hub
try:
    from .forms_registry import FORMS
except Exception:
    FORMS = []


# =========================
# Helpers
# =========================
def can_export_requests(user) -> bool:
    """
    Only managers / board / superuser can export Excel.

    Current role system:
    - MIDDLE = direct manager
    - EXECUTIVE = executive management
    - BOARD = HR / final manager
    """
    if not getattr(user, "is_authenticated", False):
        return False

    if user.is_superuser:
        return True

    profile = getattr(user, "profile", None)

    if not profile:
        return False

    return profile.role in [
        Role.MIDDLE,
        Role.EXECUTIVE,
        Role.BOARD,
    ]


def get_visible_requests_queryset(user):
    """
    Show only requests related to the logged-in user.

    Visible requests:
    - Requests created by the user
    - Requests waiting for this user as targeted approver / replacement
    - Requests already decided by this user
    - For Middle Management: same-department requests waiting for Middle approval
    - For Executive: requests waiting for Executive approval
    - For Board / HR: requests waiting for Board approval
    - Superuser: all requests
    """
    qs = (
        Request.objects
        .all()
        .select_related(
            "created_by",
            "created_by__profile",
            "required_user",
            "required_user__profile",
        )
        .prefetch_related("approvals", "leave")
    )

    if user.is_superuser:
        return qs

    user_profile = getattr(user, "profile", None)

    related_filter = (
        Q(created_by=user)
        | Q(required_user=user)
        | Q(approvals__decided_by=user)
    )

    if user_profile:
        # Middle manager can see same-department requests waiting for middle approval
        if user_profile.role == Role.MIDDLE:
            user_departments = user_profile.departments.all()

            related_filter |= Q(
                required_user__isnull=True,
                required_role=Role.MIDDLE,
                created_by__profile__departments__in=user_departments,
            )

        # Executive can see requests waiting for executive approval
        if user_profile.role == Role.EXECUTIVE:
            related_filter |= Q(
                required_user__isnull=True,
                required_role=Role.EXECUTIVE,
            )

        # Board / HR can see requests waiting for final board approval
        if user_profile.role == Role.BOARD:
            related_filter |= Q(
                required_user__isnull=True,
                required_role=Role.BOARD,
            )

    return qs.filter(related_filter).distinct()


def apply_request_filters(qs, request):
    """
    Apply request list filters.
    Used by request_list and request_export_excel.
    """
    selected_mine = request.GET.get("mine", "")
    selected_pending = request.GET.get("pending", "")
    selected_title = request.GET.get("title", "")
    selected_status = request.GET.get("status", "")
    selected_role = request.GET.get("role", "")
    selected_creator = request.GET.get("creator", "")
    selected_date = request.GET.get("date", "")

    if selected_mine == "1":
        qs = qs.filter(created_by=request.user)

    if selected_pending == "1":
        qs = qs.exclude(
            status__in=[
                RequestStatus.APPROVED,
                RequestStatus.REJECTED,
                RequestStatus.CANCELLED,
            ]
        )

    if selected_title:
        qs = qs.filter(title=selected_title)

    if selected_status:
        qs = qs.filter(status=selected_status)

    if selected_role:
        qs = qs.filter(required_role=selected_role)

    if selected_creator:
        qs = qs.filter(created_by_id=selected_creator)

    if selected_date:
        qs = qs.filter(created_at__date=selected_date)

    return qs


def user_can_view_request(user, obj: Request) -> bool:
    """
    Security check for request_detail and leave_detail.
    """
    if user.is_superuser:
        return True

    user_profile = getattr(user, "profile", None)

    is_related = (
        obj.created_by == user
        or obj.required_user == user
        or obj.approvals.filter(decided_by=user).exists()
    )

    if is_related:
        return True

    if not user_profile:
        return False

    # Middle manager can view same-department requests waiting for middle approval
    if user_profile.role == Role.MIDDLE:
        creator_profile = getattr(obj.created_by, "profile", None)

        if creator_profile:
            user_departments = user_profile.departments.all()
            creator_departments = creator_profile.departments.all()

            same_department = creator_departments.filter(
                pk__in=user_departments.values_list("pk", flat=True)
            ).exists()

            if (
                obj.required_user_id is None
                and obj.required_role == Role.MIDDLE
                and same_department
            ):
                return True

    # Executive can view requests waiting for executive approval
    if user_profile.role == Role.EXECUTIVE:
        if (
            obj.required_user_id is None
            and obj.required_role == Role.EXECUTIVE
        ):
            return True

    # Board can view requests waiting for board approval
    if user_profile.role == Role.BOARD:
        if (
            obj.required_user_id is None
            and obj.required_role == Role.BOARD
        ):
            return True

    return False


# =========================
# Requests generic
# =========================
@login_required
def request_list(request):
    user = request.user

    qs = get_visible_requests_queryset(user)

    # Keep accessible requests before table filters,
    # so dropdown options are based on visible data.
    options_qs = qs

    # Dropdown option lists
    title_options = (
        options_qs
        .exclude(title__isnull=True)
        .exclude(title__exact="")
        .values_list("title", flat=True)
        .distinct()
        .order_by("title")
    )

    role_options = [
        {"value": str(value), "label": label}
        for value, label in Role.choices
        if options_qs.filter(required_role=value).exists()
    ]

    creator_options = (
        options_qs
        .values("created_by__id", "created_by__username")
        .distinct()
        .order_by("created_by__username")
    )

    date_options = (
        options_qs
        .annotate(request_date=TruncDate("created_at"))
        .values_list("request_date", flat=True)
        .distinct()
        .order_by("-request_date")
    )

    # Selected filters
    selected_title = request.GET.get("title", "")
    selected_status = request.GET.get("status", "")
    selected_role = request.GET.get("role", "")
    selected_creator = request.GET.get("creator", "")
    selected_date = request.GET.get("date", "")
    selected_mine = request.GET.get("mine", "")
    selected_pending = request.GET.get("pending", "")

    qs = apply_request_filters(qs, request)

    return render(
        request,
        "core/request_list.html",
        {
            "requests": qs,
            "title_options": title_options,
            "role_options": role_options,
            "creator_options": creator_options,
            "date_options": date_options,
            "selected_title": selected_title,
            "selected_status": selected_status,
            "selected_role": selected_role,
            "selected_creator": selected_creator,
            "selected_date": selected_date,
            "selected_mine": selected_mine,
            "selected_pending": selected_pending,
            "can_export_excel": can_export_requests(user),
        },
    )


@login_required
def request_export_excel(request):
    """
    Export current visible / filtered request list to Excel.
    Only Middle Manager, Executive, Board / HR, and superuser can export.
    """
    if not can_export_requests(request.user):
        return HttpResponseForbidden("You are not allowed to export requests.")

    qs = get_visible_requests_queryset(request.user)
    qs = apply_request_filters(qs, request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Requests"

    headers = [
        "ID",
        "Title",
        "Status",
        "Required Role",
        "Creator",
        "Date of Request",
        "Required User",
        "Body",
    ]

    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="155F83")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9E2EC"),
        right=Side(style="thin", color="D9E2EC"),
        top=Side(style="thin", color="D9E2EC"),
        bottom=Side(style="thin", color="D9E2EC"),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for obj in qs:
        required_user = ""
        if obj.required_user:
            required_user = obj.required_user.get_full_name() or obj.required_user.username

        creator = obj.created_by.get_full_name() or obj.created_by.username

        created_at = ""
        if obj.created_at:
            created_at = timezone.localtime(obj.created_at).strftime("%Y-%m-%d %H:%M")

        ws.append(
            [
                obj.pk,
                obj.title,
                obj.status,
                obj.get_required_role_display(),
                creator,
                created_at,
                required_user,
                obj.body or "",
            ]
        )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        "A": 10,
        "B": 34,
        "C": 16,
        "D": 22,
        "E": 22,
        "F": 22,
        "G": 24,
        "H": 45,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row_number in range(1, ws.max_row + 1):
        ws.row_dimensions[row_number].height = 24

    filename = f"requests_export_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)

    return response


@login_required
def request_create(request):
    if request.method == "POST":
        form = RequestForm(request.POST)

        if form.is_valid():
            obj = form.save(commit=False)
            obj.created_by = request.user
            obj.status = RequestStatus.DRAFT
            obj.save()

            messages.success(request, "Draft created.")
            return redirect("request_detail", pk=obj.pk)

    else:
        form = RequestForm()

    return render(request, "core/request_form.html", {"form": form})


@login_required
def request_detail(request, pk):
    obj = get_object_or_404(
        Request.objects
        .select_related(
            "created_by",
            "created_by__profile",
            "required_user",
            "required_user__profile",
        )
        .prefetch_related("approvals", "leave"),
        pk=pk,
    )

    if not user_can_view_request(request.user, obj):
        return HttpResponseForbidden("You are not allowed to view this request.")

    can_act = (not obj.is_terminal()) and can_user_approve(request.user, obj)

    # For applicant, manager, executive, and HR/board, show applicant's leave stats
    stats = leave_year_stats(obj.created_by) if hasattr(obj, "leave") else None

    return render(
        request,
        "core/request_detail.html",
        {
            "obj": obj,
            "can_act": can_act,
            "stats": stats,
        },
    )


@login_required
@require_POST
def request_submit(request, pk):
    obj = get_object_or_404(Request, pk=pk, created_by=request.user)

    if obj.status != RequestStatus.DRAFT:
        messages.info(request, "Only drafts can be submitted.")
        return redirect("request_detail", pk=obj.pk)

    obj.submit()

    messages.success(request, "Request submitted for approval.")
    return redirect("request_detail", pk=obj.pk)


@login_required
@require_POST
def request_approve(request, pk):
    obj = get_object_or_404(Request, pk=pk)

    if not can_user_approve(request.user, obj):
        return HttpResponseForbidden("Not allowed")

    apply_decision(
        request.user,
        obj,
        Decision.APPROVE,
        comment=request.POST.get("comment", ""),
    )

    messages.success(request, "Approved.")
    return redirect("request_detail", pk=obj.pk)


@login_required
@require_POST
def request_reject(request, pk):
    obj = get_object_or_404(Request, pk=pk)

    if not can_user_approve(request.user, obj):
        return HttpResponseForbidden("Not allowed")

    apply_decision(
        request.user,
        obj,
        Decision.REJECT,
        comment=request.POST.get("comment", ""),
    )

    messages.warning(request, "Rejected.")
    return redirect("request_detail", pk=obj.pk)


@login_required
@require_POST
def request_cancel(request, pk):
    obj = get_object_or_404(Request, pk=pk)

    # Only applicant / creator can cancel the request
    if obj.created_by != request.user:
        return HttpResponseForbidden("Not allowed")

    # Final requests cannot be cancelled again
    if obj.status in [
        RequestStatus.APPROVED,
        RequestStatus.REJECTED,
        RequestStatus.CANCELLED,
    ]:
        messages.error(request, "This request cannot be cancelled.")
        return redirect("request_detail", pk=obj.pk)

    obj.status = RequestStatus.CANCELLED
    obj.required_user = None
    obj.save(update_fields=["status", "required_user"])

    messages.success(request, "Request cancelled successfully.")
    return redirect("request_detail", pk=obj.pk)


# =========================
# Leave typed request
# =========================
@login_required
def leave_new(request):
    """
    Create the typed LeaveRequest first, so replacement is known.
    Then call base.submit(), which sets the first approver.
    Also show the user's current-year leave balance.
    """
    stats = leave_year_stats(request.user)

    if request.method == "POST":
        form = LeaveRequestForm(request.POST, request=request)

        if form.is_valid():
            with transaction.atomic():
                base = Request.objects.create(
                    title=f"Leave Request - {request.user.get_username()}",
                    body="Leave request submitted via form",
                    created_by=request.user,
                    status=RequestStatus.DRAFT,
                )

                leave = form.save(commit=False)

                if not leave.full_name:
                    leave.full_name = (
                        request.user.get_full_name()
                        or request.user.get_username()
                    )

                leave.base_request = base
                leave.save()

                base.submit()

                messages.success(request, "Leave request submitted.")
                return redirect("request_detail", pk=base.pk)

    else:
        initial = {
            "full_name": request.user.get_full_name()
            or request.user.get_username()
        }
        form = LeaveRequestForm(initial=initial, request=request)

    return render(
        request,
        "core/leave_form.html",
        {
            "form": form,
            "stats": stats,
        },
    )


@login_required
def leave_detail(request, pk):
    base = get_object_or_404(
        Request.objects
        .select_related(
            "created_by",
            "created_by__profile",
            "required_user",
            "required_user__profile",
        )
        .prefetch_related("approvals", "leave"),
        pk=pk,
    )

    if not user_can_view_request(request.user, base):
        return HttpResponseForbidden("You are not allowed to view this leave request.")

    leave = getattr(base, "leave", None)

    # Show applicant's leave stats
    stats = leave_year_stats(base.created_by)

    return render(
        request,
        "core/leave_detail.html",
        {
            "obj": base,
            "leave": leave,
            "stats": stats,
        },
    )


# =========================
# New Request hub
# =========================
@login_required
def new_request_hub(request):
    return render(request, "core/request_new_hub.html", {"forms_list": FORMS})


# =========================
# Profile Detail
# =========================
@login_required
def profile_detail(request, user_id):
    profile = get_object_or_404(Profile, user__id=user_id)
    
    # Check if user has permission to view this profile
    if request.user.id != user_id and not request.user.is_superuser:
        # Allow managers to view their team members' profiles
        if hasattr(request.user, 'profile'):
            if request.user.profile.role not in [Role.MIDDLE, Role.EXECUTIVE, Role.BOARD]:
                return HttpResponseForbidden("You are not allowed to view this profile.")
    
    # Leave stats for the current year
    stats = leave_year_stats(profile.user)

    # All leave requests for this user
    leave_requests = LeaveRequest.objects.filter(
        base_request__created_by=profile.user
    ).select_related('base_request')

    # All approvals made by this user
    approvals = Approval.objects.filter(
        decided_by=profile.user
    ).select_related('request')

    context = {
        "profile": profile,
        "stats": stats,
        "leave_requests": leave_requests,
        "approvals": approvals,
    }
    return render(request, "core/profile_detail.html", context)


# =========================
# Logout GET confirm + POST
# =========================
@login_required
@require_http_methods(["GET", "POST"])
def logout_page(request):
    if request.method == "POST":
        logout(request)
        return redirect("login")

    return render(request, "registration/logout.html")
    
    
    # core/views.py - Add this function at the end of the file
import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import jdatetime
from datetime import datetime

@csrf_exempt
@require_http_methods(["POST"])
def convert_to_persian(request):
    """API endpoint to convert Gregorian date to Persian (Jalali) date"""
    try:
        data = json.loads(request.body)
        gregorian_date = data.get('date')
        
        if gregorian_date:
            try:
                # Parse the gregorian date
                dt = datetime.strptime(gregorian_date, '%Y-%m-%d')
                # Convert to Persian date
                jd = jdatetime.date.fromgregorian(date=dt.date())
                
                persian_months = ['فروردین', 'اردیبهشت', 'خرداد', 'تیر', 'مرداد', 'شهریور', 
                                 'مهر', 'آبان', 'آذر', 'دی', 'بهمن', 'اسفند']
                
                return JsonResponse({
                    'persian': f"{jd.year}/{jd.month:02d}/{jd.day:02d}",
                    'persian_long': f"{jd.day} {persian_months[jd.month-1]} {jd.year}",
                    'success': True
                })
            except Exception as e:
                return JsonResponse({'error': str(e), 'success': False}, status=400)
        
        return JsonResponse({'error': 'No date provided', 'success': False}, status=400)
        
    except Exception as e:
        return JsonResponse({'error': str(e), 'success': False}, status=400)
import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods

@csrf_exempt
@require_http_methods(["POST"])
def convert_to_persian(request):
    """API endpoint to convert Gregorian date to Persian (Jalali) date"""
    try:
        data = json.loads(request.body)
        gregorian_date = data.get('date')
        
        if gregorian_date:
            # Simple Persian date conversion
            try:
                from datetime import datetime
                dt = datetime.strptime(gregorian_date, '%Y-%m-%d')
                
                # Simple Persian date calculation
                # For now, return a formatted version
                persian_months = ['Farvardin', 'Ordibehesht', 'Khordad', 'Tir', 'Mordad', 
                                 'Shahrivar', 'Mehr', 'Aban', 'Azar', 'Dey', 'Bahman', 'Esfand']
                
                # This is a placeholder - for accurate conversion, install jdatetime
                return JsonResponse({
                    'persian': f"{gregorian_date} (Persian)",
                    'success': True
                })
            except Exception as e:
                return JsonResponse({'error': str(e), 'success': False}, status=400)
        
        return JsonResponse({'error': 'No date provided', 'success': False}, status=400)
        
    except Exception as e:
        return JsonResponse({'error': str(e), 'success': False}, status=400)
